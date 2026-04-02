from openpyxl import load_workbook
from typing import List, Dict, Any


def parse_xlsx(file_path: str) -> List[Dict[str, Any]]:
    """
    Parses each sheet as a section. Returns list of {page, section, text} dicts.
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    results = []
    fake_page = 1

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_text = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(cell) for cell in row if cell is not None]
            if cells:
                rows_text.append("\t".join(cells))

        if rows_text:
            results.append({
                "page": fake_page,
                "section": sheet_name,
                "text": "\n".join(rows_text),
            })
            fake_page += 1

    wb.close()
    return results
